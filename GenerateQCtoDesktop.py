import datetime
import math
import os
import sys

from openpyxl import load_workbook
from openpyxl.drawing.image import Image



from getResolve import get_resolve


if __name__ == "__main__":
    path = sys.path[0]
    wb = load_workbook(path + '/template.xlsx')

    ws = wb.active
    ##img = Image('test.png')
    ##ws.add_image(img, 'A1')

    resolve = get_resolve()
    project_manager = resolve.GetProjectManager()
    project = project_manager.GetCurrentProject()
    timeline = project.GetCurrentTimeline()
    project.SetCurrentTimeline(timeline)
    clipsV1 = timeline.GetItemListInTrack("video", 1)
    clipsV2 = timeline.GetItemListInTrack("video", 2)
    current_clip = timeline.GetCurrentVideoItem()
    audioTracks = timeline.GetTrackCount('audio')
    disk_path = ""
    sourceTC = ""
    z = 23

    lenv1 = 0
    lenv2 = 0
    size = 0
    lenAudio = 0
    videoSize = 0.00
    audioSize = 0.00
    cache = ""

    print("")
    print("-------------------ClipInfos-------------------")

    # get template


# Converts SMPTE timecode to frame count
    def getframes(tc, fps, df):

        if int(tc[9:]) > fps:
            raise ValueError('SMPTE timecode to frame rate mismatch.', tc, fps)

        hours = int(tc[:2])
        minutes = int(tc[3:5])
        seconds = int(tc[6:8])
        frames = int(tc[9:])

        totalMinutes = int(60 * hours + minutes)

        # Drop frame calculation using the Duncan/Heidelberger method.
        if df:

            dropFrames = int(round(fps * 0.066666))
            timeBase = int(round(fps))

            hourFrames = int(timeBase * 60 * 60)
            minuteFrames = int(timeBase * 60)

            frm = int(((hourFrames * hours) + (minuteFrames * minutes) + (timeBase * seconds) + frames) - (
                        dropFrames * (totalMinutes - (totalMinutes // 10))))

        # Non drop frame calculation.
        else:

            fps = int(round(fps))
            frm = int((totalMinutes * 60 + seconds) * fps + frames)

        return frm


    def getTC(fps, df, frames):
        frames = abs(frames)
 # Drop frame calculation via Duncan/Heidelberger method.
        if df:

            spacer = ':'
            spacer2 = ';'

            dropFrames = int(round(fps * .066666))
            framesPerHour = int(round(fps * 3600))
            framesPer24Hours = framesPerHour * 24
            framesPer10Minutes = int(round(fps * 600))
            framesPerMinute = int(round(fps) * 60 - dropFrames)

            frames = frames % framesPer24Hours

            d = frames // framesPer10Minutes
            m = frames % framesPer10Minutes

            if m > dropFrames:
                frames = frames + (dropFrames * 9 * d) + dropFrames * ((m - dropFrames) // framesPerMinute)

            else:
                frames = frames + dropFrames * 9 * d

            frRound = int(round(fps))
            hr = int(frames // frRound // 60 // 60)
            mn = int((frames // frRound // 60) % 60)
            sc = int((frames // frRound) % 60)
            fr = int(frames % frRound)

# Non drop frame calculation.
        else:

            fps = int(round(fps))
            spacer = ':'
            spacer2 = spacer

            frHour = fps * 3600
            frMin = fps * 60

            hr = int(frames // frHour)
            mn = int((frames - hr * frHour) // frMin)
            sc = int((frames - hr * frHour - mn * frMin) // fps)
            fr = int(round(frames - hr * frHour - mn * frMin - sc * fps))


# Return SMPTE timecode string.
        return (
                str(hr).zfill(2) + spacer +
                str(mn).zfill(2) + spacer +
                str(sc).zfill(2) + spacer2 +
                str(fr).zfill(2)
                )



# run through the layer 1 grabbing info and markers
    for i in clipsV1:
        lenv1 += i.GetDuration()
        mediapool_item = i.GetMediaPoolItem()
        disk_path = mediapool_item.GetClipProperty('File Path')
        clipProperty = mediapool_item.GetClipProperty();
        file_size = os.path.getsize(disk_path)
        videoSize += file_size
        dateCreated = mediapool_item.GetClipProperty('Date Created')
# get clipMarkers + values
        if len(i.GetMarkers()) > 0:
            #print(i.GetMarkers())
            note = i.GetMarkers().values()
            key = str(i.GetMarkers().keys())
            my_list = list(i.GetMarkers().values())
# get marker TC in frames
            for i in key:
                if(i.isdigit()):
                    sourceTC += i;

            print ('File Name:', mediapool_item.GetClipProperty('File Name'),
                   '\tscene: ', mediapool_item.GetClipProperty('Scene'), '/', mediapool_item.GetClipProperty('Take'),
                   '\tsourceTC: ',  getTC(25, False, int(sourceTC) + getframes(mediapool_item.GetClipProperty('Start TC'), 25, False)),
                   '\tcomments: ', my_list[0]["note"], )


            scene = (mediapool_item.GetClipProperty('Scene') + '/' + mediapool_item.GetClipProperty('Take'))

            ws['C' + str(z)].value = scene
            ws['d' + str(z)].value = mediapool_item.GetClipProperty('File Name')
            ws['h' + str(z)].value = (my_list[0]["note"])
            ws['l' + str(z)].value = getTC(25, False, int(sourceTC) + getframes(mediapool_item.GetClipProperty('Start TC'), 25, False))

            sourceTC = ""
            z = z +1


# layer 2 duration
    for i in clipsV2:
        lenv2 += i.GetDuration()


# get durations
    lenv1 = datetime.timedelta(seconds=math.floor(lenv1/25))
    lenv2 = datetime.timedelta(seconds=math.floor(lenv2/25))
    lenAudio = datetime.timedelta(seconds=math.floor(lenAudio/25))


# going folders up searching for 'Drehtag'
    recordingDay = os.path.basename(os.path.dirname(os.path.dirname(os.path.dirname(disk_path))))


# get Audio Size
    Folderpath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(disk_path)))))
    Folderpath = Folderpath + '/_______A' + '/' + recordingDay

    for path, dirs, files in os.walk(Folderpath):
        for f in files:
            fp = os.path.join(path, f)
            size += os.path.getsize(fp)


#output
    print ("")
    print ("-------------------Allgemeine Infos-------------------")
    print('Drehdatum: ', dateCreated)
    print ('Drehtag: ', recordingDay)
    print('')
    print(len(clipsV1), '     Kopierer Clips ', lenv1, ' Dauer')
    #print(len(clipsV2), 'Nichtkopierer Clips ', lenv2, ' Dauer')
    print(len(clipsV2 + clipsV1), '  Gesamt Alle Clips ', lenv1 + lenv2, ' Dauer')
    print(" ")
    print("VideoMaterial :", round((videoSize / 1024 / 1024 / 1024), 2), "GB")
    print("AudioMaterial :", round((size / 1024 / 1024 / 1024), 2), "GB")
    print('AudioTracks: ', audioTracks)


    ws['L12'].value = recordingDay
    ws['L14'].value = dateCreated
    ws['E53'].value = lenv1 + lenv2
    ws['E54'].value = lenv1
    ws['E56'].value = round((videoSize / 1024 / 1024 / 1024), 2) + round((size / 1024 / 1024 / 1024), 2)
    ws['L53'].value = round((videoSize / 1024 / 1024 / 1024), 2)
    ws['L54'].value = round((size / 1024 / 1024 / 1024), 2)



    wb.save(filename= os.path.expanduser("~/Desktop/") + recordingDay + '.xlsx', )
